"""
Camp Carysbrook Day-Off Scheduler — Streamlit App
Version 2
2nd try
"""

import io
import pandas as pd
import streamlit as st
import pulp
from collections import defaultdict

st.set_page_config(
    page_title="Camp Carysbrook · Staff Scheduler",
    page_icon="🌲",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Carysbrook Brand CSS ──────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=Lato:wght@300;400;600&display=swap');

  html, body, [class*="css"] {
      font-family: 'Lato', sans-serif;
      background-color: #F5F0E8;
      color: #2E2E2E;
  }

  #MainMenu {visibility: visible;}
  footer {visibility: hidden;}
  header {visibility: visible;}

  /* ── Header banner ── */
  .cc-header {
      background: linear-gradient(135deg, #2D5A27 60%, #3D7035);
      padding: 1.6rem 2rem;
      border-radius: 10px;
      margin-bottom: 1.8rem;
      display: flex;
      align-items: center;
      gap: 1.4rem;
      box-shadow: 0 3px 12px rgba(45,90,39,0.25);
  }
  .cc-header-text h1 {
      font-family: 'Playfair Display', serif;
      color: #FFFFFF;
      font-size: 1.9rem;
      margin: 0;
      line-height: 1.2;
  }
  .cc-header-text p {
      color: #C8DFC5;
      font-size: 0.85rem;
      margin: 0.25rem 0 0 0;
      font-weight: 300;
      letter-spacing: 0.06em;
      text-transform: uppercase;
  }
  .cc-icon { font-size: 2.8rem; }

  /* ── Headings ── */
  h2, h3 {
      font-family: 'Playfair Display', serif !important;
      color: #5C3317 !important;
  }

  /* ── Primary button ── */
  .stButton > button[kind="primary"] {
      background-color: #5C3317 !important;
      border: none !important;
      color: white !important;
      font-family: 'Lato', sans-serif !important;
      font-weight: 600 !important;
      letter-spacing: 0.06em !important;
      border-radius: 6px !important;
      padding: 0.65rem 1.6rem !important;
      transition: background-color 0.2s ease !important;
  }
  .stButton > button[kind="primary"]:hover {
      background-color: #7A4420 !important;
  }

  /* ── Download buttons ── */
  .stDownloadButton > button {
      background-color: #FFFFFF !important;
      border: 1.5px solid #2D5A27 !important;
      color: #2D5A27 !important;
      font-family: 'Lato', sans-serif !important;
      font-weight: 600 !important;
      border-radius: 6px !important;
      transition: all 0.2s ease !important;
  }
  .stDownloadButton > button:hover {
      background-color: #2D5A27 !important;
      color: white !important;
  }

  /* ── Sidebar ── */
  [data-testid="stSidebar"] {
      background-color: #2D5A27 !important;
  }
  [data-testid="stSidebar"] * {
      color: #E8F0E7 !important;
  }
  [data-testid="stSidebar"] h1,
  [data-testid="stSidebar"] h2,
  [data-testid="stSidebar"] h3 {
      color: #FFFFFF !important;
      font-family: 'Playfair Display', serif !important;
  }
  [data-testid="stSidebar"] hr {
      border-color: #4A7A43 !important;
  }
  [data-testid="stSidebar"] code {
      background-color: #1F3F1C !important;
      color: #C8DFC5 !important;
      border-radius: 4px;
      padding: 0.2rem 0.4rem;
  }
  [data-testid="stSidebar"] .stDownloadButton > button {
      background-color: #5C3317 !important;
      border: none !important;
      color: white !important;
  }
  [data-testid="stSidebar"] .stDownloadButton > button:hover {
      background-color: #7A4420 !important;
  }

  /* ── File uploader ── */
  [data-testid="stFileUploader"] {
      background-color: #FFFFFF;
      border: 2px dashed #2D5A27 !important;
      border-radius: 8px;
  }

  /* ── Expander ── */
  [data-testid="stExpander"] {
      border: 1px solid #C8DFC5 !important;
      border-radius: 8px !important;
      background-color: #FFFFFF;
  }

  /* ── Alerts ── */
  [data-testid="stAlert"] {
      border-radius: 6px !important;
  }

  /* ── Cards / metric boxes ── */
  [data-testid="stMetric"] {
      background-color: #FFFFFF;
      border: 1px solid #C8DFC5;
      border-left: 4px solid #5C3317;
      border-radius: 8px;
      padding: 0.8rem 1rem;
  }

  /* ── Footer ── */
  .cc-footer {
      margin-top: 3rem;
      padding-top: 1rem;
      border-top: 1px solid #C8B89A;
      text-align: center;
      color: #9A7B5A;
      font-size: 0.78rem;
      letter-spacing: 0.05em;
  }

  /* ── Bar chart override colour ── */
  [data-testid="stVegaLiteChart"] canvas { border-radius: 6px; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="cc-header">
  <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAHgAAAB4CAYAAAA5ZDbSAAAAGXRFWHRTb2Z0d2FyZQBBZG9iZSBJbWFnZVJlYWR5ccllPAAAAyhpVFh0WE1MOmNvbS5hZG9iZS54bXAAAAAAADw/eHBhY2tldCBiZWdpbj0i77u/IiBpZD0iVzVNME1wQ2VoaUh6cmVTek5UY3prYzlkIj8+IDx4OnhtcG1ldGEgeG1sbnM6eD0iYWRvYmU6bnM6bWV0YS8iIHg6eG1wdGs9IkFkb2JlIFhNUCBDb3JlIDUuNi1jMTQ1IDc5LjE2MzQ5OSwgMjAxOC8wOC8xMy0xNjo0MDoyMiAgICAgICAgIj4gPHJkZjpSREYgeG1sbnM6cmRmPSJodHRwOi8vd3d3LnczLm9yZy8xOTk5LzAyLzIyLXJkZi1zeW50YXgtbnMjIj4gPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9IiIgeG1sbnM6eG1wPSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvIiB4bWxuczp4bXBNTT0iaHR0cDovL25zLmFkb2JlLmNvbS94YXAvMS4wL21tLyIgeG1sbnM6c3RSZWY9Imh0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC9zVHlwZS9SZXNvdXJjZVJlZiMiIHhtcDpDcmVhdG9yVG9vbD0iQWRvYmUgUGhvdG9zaG9wIENDIDIwMTkgKE1hY2ludG9zaCkiIHhtcE1NOkluc3RhbmNlSUQ9InhtcC5paWQ6MUFBNTM1N0RCQUMxMTFFOTk5MTc4RERDMTE2Q0Y4MUUiIHhtcE1NOkRvY3VtZW50SUQ9InhtcC5kaWQ6MUFBNTM1N0VCQUMxMTFFOTk5MTc4RERDMTE2Q0Y4MUUiPiA8eG1wTU06RGVyaXZlZEZyb20gc3RSZWY6aW5zdGFuY2VJRD0ieG1wLmlpZDoxQUE1MzU3QkJBQzExMUU5OTkxNzhEREMxMTZDRjgxRSIgc3RSZWY6ZG9jdW1lbnRJRD0ieG1wLmRpZDoxQUE1MzU3Q0JBQzExMUU5OTkxNzhEREMxMTZDRjgxRSIvPiA8L3JkZjpEZXNjcmlwdGlvbj4gPC9yZGY6UkRGPiA8L3g6eG1wbWV0YT4gPD94cGFja2V0IGVuZD0iciI/PhikGEAAABo8SURBVHja7F0LvE1VGt/3uvKIdPXgFlHKxAgzmtFDUhRRkSaTW0opScmzUMojjaYoMxkTSd6aTCHPHhPh5hGuUEyEiSlJJuT9mu/f/e8s6659zzn7rH3O5e71+32/fc7ea++99vrW+t7rWynHjh1zwnLqltSwC0IEhyVEcFhCBIclRHBYQgSHJURwWEIEF6CSFusNKSkp+fJDGj50d1k5VBG4WKAc4SyBdIEiAiVZ9bDAXoE9Aj8IfC+wWeBrgbUCX743bNye/PiNfoxSKbHelF8QLAj9lRwaCdQVuFqgjK1+FFgtkCXwkcCHgvD/hQhODFJryeEOwkUJeu1RInuCwDuC7G0hgu0i9Qw53CvQVqBakicRyPu7AsMEPhBkHwsR7B+xGXLoKvCgwBn5kCV+LvCiwHhB9OEQwdEjFoJRL4F2AkVjvB1C02cC/xbYJLCF534UOKDUK0HhC4JZeYGKApdRSCsS4zu/EugnME4QfTREsDdiIdE/KtBb4Mwob/uvwGwKQ1nSwf+x0IbfUGi7QeA6gWJR3v6pQEdpw8IQwWbhaYRAzSiqY1aOF5gknbksYGoC5DYWaCHQTOC0KG57VaC7tG1XgUcwZ0wvQqEI1T8QeEVgpnTekSRI8GdT2HtM4III1aFXt5J2flxgESwddj5mocCVEapOAdmWzlrp4x3gq5UMeu5n8rxNcQxKqGnPGp6tq1d/Euhja0CeNAiWTrpWDm8JnJtHtfkCnaMhw/K8wlLvkOH8YPBFD2NGD7nnBdbDR4F3LuaAmh9JMiai7xfoL3BOHlXfF2gpz9tRIBAsHQMy95pAYY8q2ymsTIjiWbfKoQOpQB25Z4V2/fdyeIlC00R2NoSnTAHMqnJyz1apV0F+LxU4W1GBWgpUFnhP6vyURxtKyWEApX6vzoFU31iesyHRCE6os0E6o4ccRuWBXJDsS6NBLktzgQYCp0NNkeefoFbJc5bIYT3/LpH/ePfD/A+eX5r1IH0/wPO7Bd4kwv6JASfPnc7BkqvIvTsF2stPUKWNHu2EWXWRPKNGomWH1AQi93mOdFPZL9BGOqqFwA8+Ho8B8WvyPK/SSNowgLMY5UuCSkr3UFIG6c7meejFTZwIjhlp93xqAW95VAEZnydtqH3KIVg+CoaA7h6XIexcKR00Mo/7LxW4xHDJ1ZehP38j0EnqXe/1GPBdgdo0ejRT+az83ocDEYrBOFSlKpiBbEsZgVvQJgOSdwn8UX52oZClF1jk3qdaeGogWD4GH/t0HsaBK3TeyftqCEwSAG/E9YkUbEzlcs5G8MDRUs9kKHmW9ZYRifcZ6kzmsTN5dDtSlX/DSiXPBf8GH4U9eo38nyFQ2oDol+VwCymCCcmzTQPkpEOwfMQf5DDQ4zKsT/WkM74z3FefMwb31yJCdji5bdKnK+S1Hmcx/MB/M7xvByVyd2beJO85XavjGifgG/6d1B+mXR8iUJx8/T80gkz1INkz5VCf7dYLhLmZ8v5zT1oES+Nh8hvjIVlOQwdLJ+z1UD/ecHLs0JCAy5K/NjSoGq6w9pAA9OqqTo7tOVOes4x8P5N1Bsj/OqQGKNVoEXPfW4bWKpTh8q7VhnZfzGMH3g9TaR1+qwnJiznwthsuXyjwjtx72kmHYJKtdxyzHXeOwJ3y8QfzILflSXK7YYYLfOHhmnM7B3Uwe8H/zuK5iuS1nwh8TB0X6s5a/gesZntBIdYJtHYtUXIuVfmeuhx4H/LUnyl0uTz8tjyEr1Wc6SbTJdS3QUEiOC2g577ODtbLSgo3e/O41yVb23WkUtCCqRBkfRbVj5+tXXINM2oGWQIoRFYeFqR6Br44mXwTXqZhVJH2ynOry/FfAqvIm38nUJ11IF8MpiCG9kFShrOkq7z7gILkT+Vac7ZZVxEflWsfSZ3JJ8UMlsY+oJA6tSAS4uYojPCuLlnDIMA0o8QMRB+iFD0PM13ge3n2fwUeF5hnQi7IocDdhlmGe++lKgMPUgdlEO7hRKhCucAl+XBAIJTnIs5SlMcFHoHaJu8ppL0Dg6S9xzePoOk2fyNYGlmOfFMv6GxIo5sjPYOdtYoC1Fh5ZgnFYuSS0GW0TZ8jx2sFBkUZSnMJVR6vdx8RmCswQjn3FSlGUZLi5dTbUSq5qhbJ/MWKAWao4fl47nDDq0t7CIb5bgajkSUN53vH6FmB0LSPvGsDLEk091Ulr3ydHRarfRcDzE8sV28eJyrIPkC92R18szkA9nGAtpXzTxme1ZHX9dJU6t+ebxEsjYMh4VbDpSwaDqIudJrXo034HFqSypAc3+A3rJXsoYKP+8Bv73JyojccqmM/UyTy3Tls7z6aTkHm4bfuTTej+qz9JPMmIXMQfdDWihVnAyXMVQbyh1FeA4YCnRc6OeGuGGBz85CoHQpP52EmS731PgcfyOdltC9fA7XLbxyVPKskpfFWHHywbVfSqM+bFCgr8PtXGp4D408/wyuelvr9Te9OmjdJGtuauqte4Avtq9WtRHLm8ivEUDWQetsD1MkLU7quyYEF/fUJgel+Ii8Y6bneOe4mxG8EB46jPn6YghmC6surErXWppWGSbETGojc86MNBKda6Dw3KkMvsAa9YPiot4ncDeyIGh4j2VqhrxhteZWnQD5LEPyS+iacyT+bNiGcYUDzfyFaupqbkKu0qbPhUilK4/mGB7dwzJENvWnAV0s3IvQ1uYZ7XJWlaoCz9w6Bf5Cfu25IOBz6U3r1O2g+pVAFRI2Sd3SgXozymFzHLFzANsBZAvtzM+0Zs6lj66W9qz3kBwR3MZxbAxXHYKRwpdH98r+4YonKDnACz6QEDoS0VaxpkMw7Szt6xoHkDyl8QR//K3VlsJp3taowyDSk0KXzOBP1w/PuT7qQJY2FVWeJodr98vFvaAhGR9+knIJEWpxqVRWpvyXAWVyWHQwEX0WjS22/cVmG57tLamAceR0GF56/gmS7oVL9Vrk+Tbt/jsG6tkbqVU02D37AcG6roxjxVZMcBZsbaK0Cic6gPXgvB0sQyG3M2TuKyHXNoUvpfLDB4+Gl6kmh6XZ5bj0BeLgWErkg46Np8DHpxiZ7dBV5xjVJm8FUPbY6uYPUn5MP7hWh04tTsOpEgWQH+WEuqTsOxMIDdSNnEKTeGexkRFzs4rmjRMKjJp90jO9L5WCFVKyT4Vry/OVSB4aSOwWul/9ztHsx6PVw3GFSr12yZvCNjnkFwsgoRvxegW4UdrKJ3AO0FNkqg8jzYUGDDfxmqi0oLSkcTiR/zJZO/tRkp45hFh/loE2hEeMvTs7aKkcRJr/gsYHhXlO/Nddt2rGWeLxJTQ3nFquRgxQoMGOwEm+4ibRJHQgk8Kf2k/9fWlSNGintqCDQXpHW99AhP4kdiIGe6qXSxFDwraBAxagxlOZA60yfsRsDblpvPF5Rs9wCPftqWvASjuBGhnOTtP/QdxGV8VuK/Qi/WeaaGrmIezhn78CAePDDfPZS6pgoz8v5e6Qd6+h1OmJpUB1lFMtGsh7XRHtQEaLWmWYrrHRy7won9/KdmxKOYGlIVVps9DJd+385jxcpggQ6YS2lb5CqIvJx5waA2PoccPDvfkOy3JJ6OGbWSrZjrLz/JYuUY5U8twp1bhdZt1IlRH/PNtgH3DLDgOAGFOASOoNN0ucm3ebs5DjH3QKnQxF+QFWFXM6yjFjYgpsrfA6diXVC39AxgJJJFgNdEwb+L2h0sFXq8DthqbtPnv1etOPDIGX/BtQvr+D7IISsKwzn5hvOYWZMUZA9lTMKwtVgnl9qefKuoqEFOi+cFBWkcz7itXQeJ7Atv+Wsvks68UqLbYC0vpmUYwoHXjVYsuB1M0VisixxTlzH7Jo9L0+omiQNzDaQkoelI1/1mFWYUUOo98K50JrWJXTq1369RDHO7KK0YGFwIkAecV8jEP0R0PtAodaTUsHRr8ZC76eU/ZQeeSL3feLkXpAHW/dgP2pSmo+Gp9Akp5fsPPgSogdhc0Ww2kM0PDxHnflQAJ2LNlaknICBCJtxNed4vFcWpejACgIFaZ79mDr3XvZRcWoN3cmXdVPvCgOCfS958UOiiznmNAefR/jgnbRk9STZ6e1h8YoHsdAb7yOJ3EC28Qp104HO8VAbqEcIo703Xj0zQnmOyIWcgYVuWCAHtnAtZYOONKPqLEYvlRIpZJ1uOPedLgRIw88j76jJEVjTyR0us9Vyh4Ln3aGocJsp3AHpcFW+oPC6Hvz9mLT1PWoASywnVKlHa9k9aq4tsAV5J3RmBPphUZvqnPjK8JyLEolgU2IU06q6qZpwcJR8dxlhKY82ySKiHrGKD/wVYTOzEDHJAaeqddeTfF/DAXEVqcpiywMOgwosaKfhWgmlX07QRgx1z0skgk3keZs2e8FbqvMvPnAyBY0F0AHlOkbkT0GlDGT05hilPaWcE6M94dqDz3aiZfOoXpZSXYOb8Gk3zlt+Z1IugAFkkXbPtybZlnJFzFJWqs9RqZfvDULBaUp92H3hXdktDYUUDT/qd+z4oKVn8LuFzokBgfBswf6cybRNQZU+nKHQbdfKuxCg/wVlD/T9AD1USf7v5qSwUtIs3fOTQRDD6B3l5OSoqkXdt6Yys9dT8AoCqefTxHcVVbIUdlphtqcUByE6egft1G/Zzlwnz8uSZ7egiliZ4BpfEFHitV56r2JW/WUW+5nBfhCcEsWHzdOsWOPZ8YUoeHxIVck2YuGqG0r+q5blFKzakc+OpnR9Fs2WiIJ8FlYw1Y1nCclvY5kpJecM8uO5ERKcHk7mDI7nY4/Ix0J1wfqispb1XkjIHdk5mKUghYgWmU3DxjOukQG8V+6ZK7//zk6HJAuddSSsTXL9M8vfvZ9myGhLoWQi+FiUqpNXgZEEUYPnYeW/BRcdOvCYPAtIhGdqi8mYL9dV8od7IMw0UwbIPeTTTyBIT66/a2nwYTkPYqBhzUMQPJa3YmFbX7bBy9aQNAQfjlJ10j8Uyjoc4i0VMl/FOb5eN14kg+T9z8NECbWosatyYOG3KsGT944m2GQZF1JKVr1lZ9Gah9Ce6/R1yBxsRaKcWIFI0aZVCOl5fGSGANYswQGQSWOEm6gk0DQGXIUwhzp5RZ4GL8yiOzHoMpDInU/zYzrfv4iz+XXDPaVtNiBmZ0Ojdq1gRPhaO43McTW1zj2TpLgjSfhW8l0IN09CjbKRHMw0oCjI/JGSaknqnJjBXQymxL4B2cOhJu6impShagxMHQFfNMj3paqbleuRdRkAqx7TEhWTZXJWX2g414mIhIAB+3MlWmmq06xZlYnM4u3IEgI3CwwRQMjPeqptIzlL2nNtrpsLZCH5YTb1U/iJ+xjileMtRUlqd+rqINmDm5U2IwqzpO+QYj8INq3OP8Pg41ypmOSgi75N4aE6VhvIEasNbKyJBaXAmqBHKAnDW7OHgswF0pnjWM9d5VeZq+kbU57AeTg+JlmWnDF7sXylLL/X0Sxrdfl3YxSTZVMiEbzPA8n68pNPaB78gO+Bya4M+RB4MOyr5TxSHsXSkVuoW2PAIFUSwlLT5fwdbgYfCjuXuaZThvxuI4l2pfjbA+DLbpgS1DJk4msj0Ik2APTFbEOOa5Mr1ncwYpqPDoVK8rlmyHAoRCxQ6oHndlVGbRpH568405pS4ADPmRcnklcQseosAbX4NWdqb2Uwl+UMxz2wSSMorzvbhZSFGBjTLSF4CAVJWMruIrgFPLiN4R5TZMnKhCGYZYUBwbUjIAHkcB0BCEDYTNF4s7RHsLj9gSbL3TxXirx4A2cFBD3kxxhFgQe+2/qwT/tMqagOZiB1pDznEfk/nhpEJcoHsMuP0W0AlPpNC/GyE41g+Fgf1M5dBUElBnsujBJp/PiJAdiBwUbg8+1BkrxRQfC1quTMd7v5Qz638HrIHH8hf39Gnv8JWVakcoWBbaKdy5KBYL1AGkQoyvIYZlhjAnho24B0YfC6WYq0CpaA0NbRVEeybDo9yBr68O/dlOwnRLlhRxPDuaU0dSZMyHIYJGdy8jf1+GiEfiIu62LlNHTSJyiwPSjXbgoAuVjr8ycOPLdgBiMeCysQ4AT4UeodoFvRRsGCvPP5XZA5sIx2BVQ5ZQB4FVP/vR9PY+JZmzQzWgQ7OW6x24jQXwaJwIvKuTstIRU6LZwaC2g9g8uwkaIHw9nQRaMYMEq8Kfd0FegWZxOQ7LwDeTpW8P9ACX4aHS1T4ejQfeHMJV3R8LxZyUKwKTMbkpeZdk5xDfemGCx3QVaZOJBaWeAVgeWKxFyLxo5aDDx3ydwuDjAIVshsB/s4oj8mkHfuiKMdDSi4DaH+f5CDbBwldwTEY/lsYQNbaGV45LdOnG7VeNyFc52cSA59v4K2Tu6Mbm6WWFM0iKs6rPNjxSJyGjvHDfSHaeTo5i7EZnFJ41GFikAdmm6JcmRwIGOtMwQ2daCPIUtazpndQ7u3iHM8yZtaJsa72ZZvBNO3i1Gvb3oBweJxLd5qNsk08jIuJNkpQxMmdEFIsSN8tAEqR3N2UDpn7vcetuWgw4Oe4SAqxn79B408r5FNjKK+X5LZ4dUCS5cpdGhCvI2K1+H/mgHBJalCDVYNEYKEl8mTTPmVu8fjZKc+GSkEd78T+7Z1sRSoV3AatKNkfoizswtJ82E5d5dHHHYHw7nlNjb+iiuFgzQAH2Vak9TdIC3CqoVgN3eZyhGqW00obJ3UhXy3CnNdHtJUSmgPSyBIGZaqQDA1bfgx1Ea7bITsADl6LgmYAx/SZjHIMJwLfyPvPGQjmiOfIdlkrFlEilZDsai5yE1xci/6dlW5cTbaZCON0gxKinp5Uj4g3Yt3JgG5R5KEdzcCZphh36QWjnm/xsG2+ifVwqiFlGdKugLpun8+mmC7kzSrx9BS11+bvdAsXjbcAs1kiK3328o2O9lDX3tYPuRyp4AXpFI2bD4C12aGofqzfhd7B4Zg8p5OhkvgMW/oO5IV9CL9cTUFTr3Ahfh3m+9KtThK4S0Za7iEdbkvhWj9BbmIfJno0fcdbG8XbzvjO/Tc7R6kulmS+7Y4j3uSiFxQNERSljdcHs3cl06+RTCd5O08Lo9NxuaM7NhyznGT6qEkDjIIo6aBDpNqlyBeaH3XFazFcczxvtB9p9Fmm+hSJsnUA4MMzgRTXmzIL62CCCEOBMEsMF+adg4DaZqeiGWjWtmSZOTW9xj0KP1sL3gLHMF0NMD/a4qUQI6K9xOM5IxkGTy4GyrSB5s8aTAS9Q3y/YHtXcioj+YePO/3RPLZSZhQuxKIXAQawB1piuKAcyXTdixawhBMJCMBWRuPy0Dy4kRts6qU0xKEXPjF3/VALljGzX42BMlXCCaSoRt7bemGZRqLErAQrJzy+/yAEVsI0SVOzt6GJrKMKI3rgsxwn1AEE8mwzjzicbkUyXXfPDaAjreoscYHA0QugvywDupRjypQh+onIrNfQhFMJA8luT7q0Q5ERMxjBh7bRc0mEFReEIQeYQWCV3QmPEnYEHtNIvlRaiJfJh+HIDh1vyG9YNnGauz5xzAcWyU9KCkaAxLZc5wc/62XZoC8IFcmcuYmBcFEMuKzEKP0lUcVCCX9iegmll5bV/ld0RJiSzBtBKJCm+ZRdSx57jYnCcXK1nY+OwirCkdF6ByUFdQVp/pRKeQ9l1ElSVFmcDrzUfltdwcac/LKsQWHfReyJislaXsXxolohPZgmWWkRC6wjGE901jTvn4ez8ZiL6RI0n3SsCohMcrOCLuRq8+CVwwJxBF+E2lXslXUcVfb7KuTEsHsvAupVtwQRXWsT4a9+x1I38r+D1BJYDXCMhVEMWKJauUonocc10+bNsekVIzgeCQzrRXFsyChY83xgCDSQpy0CFY6FHk1kFM62j1+QQbnOjkuQCxnLe/z1Rg0cLYjbnkj3w8rVJUYngGLVVebO8eccggmkotRZ+4RgcfllwIJuVcQvtxTEsGqlOrkhN52jGNmBllg0Hgxhg03QgR7IDqNkjaQ3cCJIldmgGUn1Z4RttMdFlgEa8hG4hZkyruNRpHUBCEVujuy8ExPZrD+KY9gDdnplLphxECUItSYNEsIBV/FkhM44hfaDoQLEewP4XADwqmAbDmVKAlDSEMslmtCdHMuA2EIkUEMGbw7m5wcy9rqAJPChAgOi30Ep4bddmqXfIXgoFdA2H6+GlemL5dV/4N90NKW8JIUEi0fi8QrrZ2cNPuNqP4gCgJuRKxxwubTyGhbh3VwHplr4W67W/jkNG7PjhXwc8lrkWgMu4xNket9+B4Y+iF5ww/9FnkunAzDKRU3487ccPfBvo3MdzMpNYOfP8d2YdUGFrbDBLmAz0YQwyXyv4H8xn1YTIYYM+yFhOzzsK0jfzXygPWkAAiDiO+cGycTiUZqINiRgWRYrLDHwpvy/xYKPXD+I8xnOQcC0jsgzUEvItJhR2MVPDK1Z1NtmuIc3wzTfQ/+Y4km3HowhSKtwhre6+b0msKBgf9r2S5sO/AZ29CO9w5Sok6wHOcnZqc7k/cg91ZDgbflP7IGVXRyojjwGysJ6xUkEn0jUggy2w1mIjLlwRa8Tdkt1F1LvIAIhp23Gq1cKJcyoRmQvCiPd2FmwxGA/CDnMNjtAryfaQwfUOqezXMHpd4mtR1Ojp26LP+v43E370HOy7YcODOReJSDeCMR/SKpTYFB8M8zmDtmY9QjSQuy5fQxkPQKJK3FqJu6PmTMtntJcutq96h5t8ACppPkZ9Kni11PW3PmqQlgtvPcMS55cZ9XnIgy5QKB4QV+5m/kXuyPhIwHteX3y0z+BhUMKxsSvj4rv/BgIHYI9dFvpWO6cr+hgUQK8j4igVlLbj4Jsv4kyd5H5LN38vfX5H3wBC3lNZDYf5KHA7n9SJaj5cFZ1KWfd7ejlXtu5MBA+zbznjaIlpRrW9mWH/nenhwAT8WTWCUhenBYQjUpLCGCwxIiOCwhgsMSIjhEcNgFIYLDEiI4LPm1/F+AAQA0UX8Av8SQZAAAAABJRU5ErkJggg==" style="height:70px; width:auto;" />
  <div class="cc-header-text">
    <h1>Camp Carysbrook · Staff Scheduler</h1>
    <p>Days-off scheduling &nbsp;·&nbsp; Est. 1923 &nbsp;·&nbsp; Riner, Virginia</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Colour palette for days — green → brown tones ─────────────────────────────
DAY_COLOURS = [
    "#2D5A27", "#5C3317", "#6B9E64", "#7A4420", "#4A7A43",
    "#9A6040", "#8FB887", "#C8936A", "#355E2E", "#B87040",
    "#A3C49B", "#D4956A", "#527A4C", "#E8B088",
]

REQUIRED_COLS = {"Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age"}

TEMPLATE_DATA = [
    ("Counselor 19","Lark","RRG","Archery","Riflery",25),
    ("Counselor 22","Mockingbird","","Dock","",25),
    ("Counselor 3","Sparrow","","Barn","",23),
    ("Counselor 10","Swan","","OLS","",22),
    ("Counselor 12","Hummingbird","AHC","Riflery","",21),
    ("Counselor 2","Cardinal","","Fencing","",21),
    ("Counselor 20","Sparrow","","Dock","",21),
    ("Counselor 21","Bobwhite","","Marketing","",21),
    ("Counselor 33","Swallow","AAC","Trips","",21),
    ("Counselor 16","Finch","","Barn","",20),
    ("Counselor 24","Cardinal","","Dock","",20),
    ("Counselor 27","Mallard","AHC","Dock","",20),
    ("Counselor 34","Wren","","A&C","Riflery",20),
    ("Counselor 1","Bobwhite","","A&C","",17),
    ("Counselor 13","Lark","","Nature","",19),
    ("Counselor 15","Finch","","Dock","Rec Sports",19),
    ("Counselor 23","Wren","","Store","Expeditions",19),
    ("Counselor 29","Swallow","","Barn","",19),
    ("Counselor 30","Mockingbird","AHC","Drama","",19),
    ("Counselor 32","Bluebird","G2G","Tumbling","",19),
    ("Counselor 5","Bluebird","","Climbing","",19),
    ("Counselor 6","Dove","","Riflery","Synchro",19),
    ("Counselor 9","Dove","AAC","Rec Sports","",19),
    ("Counselor 28","Oriole","AHC","Dance","",18),
    ("Counselor 31","Swan","","Canoeing","",18),
    ("Counselor 4","Nurse","","Nurse","",18),
    ("Counselor 7","Oriole","","Riflery","",18),
    ("Counselor 26","Mallard","","Archery","",17),
    ("Counselor 8","Chick","HC","Performances","",17),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def normalise(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in ["Activity1", "Activity2"]:
        df[col] = df[col].fillna("").str.strip().str.lower()
    for col in ["Cabin", "Leadership", "Name"]:
        df[col] = df[col].fillna("").str.strip()
    df["Age"] = pd.to_numeric(df["Age"], errors="coerce").fillna(0).astype(int)
    return df


def schedule(df: pd.DataFrame, num_days: int, days_per: int, max_per_day: int = None) -> pd.DataFrame:
    counselors = df["Name"].tolist()
    days = list(range(1, num_days + 1))
    n = len(counselors)

    prob = pulp.LpProblem("CampDayOff", pulp.LpMinimize)
    x = {(i, d): pulp.LpVariable(f"x_{i}_{d}", cat="Binary") for i in range(n) for d in days}

    for i in range(n):
        prob += pulp.lpSum(x[i, d] for d in days) == days_per
        
    # Daily cap: no more than max_per_day counselors off on any single day
    if max_per_day is not None:
        for d in days:
            prob += pulp.lpSum(x[i, d] for i in range(n)) <= max_per_day
            
    cabin_groups = defaultdict(list)
    for i, row in df.iterrows():
        if row["Cabin"]:
            cabin_groups[row["Cabin"]].append(i)
    for cabin, members in cabin_groups.items():
        if len(members) > 1:
            for d in days:
                prob += pulp.lpSum(x[idx, d] for idx in members) <= 1

    act_groups = defaultdict(set)
    for i, row in df.iterrows():
        for col in ["Activity1", "Activity2"]:
            if row[col]:
                act_groups[row[col]].add(i)
    for act, members in act_groups.items():
        members = list(members)
        if len(members) > 1:
            cap = max(1, len(members) // 2)
            for d in days:
                prob += pulp.lpSum(x[idx, d] for idx in members) <= cap

    lead_groups = defaultdict(list)
    for i, row in df.iterrows():
        if row["Leadership"]:
            lead_groups[row["Leadership"]].append(i)
    for role, members in lead_groups.items():
        if len(members) > 1:
            cap = max(1, len(members) // 2)
            for d in days:
                prob += pulp.lpSum(x[idx, d] for idx in members) <= cap

    ages = df["Age"].tolist()
    max_age = max(ages) if ages else 0
    seniors = [i for i in range(n) if ages[i] >= max_age - 2]
    if seniors:
        peak = pulp.LpVariable("peak_seniors", lowBound=0)
        for d in days:
            prob += pulp.lpSum(x[i, d] for i in seniors) <= peak
        prob += peak

    solver = pulp.PULP_CBC_CMD(msg=0, timeLimit=60)
    prob.solve(solver)

    status = pulp.LpStatus[prob.status]
    if status not in ("Optimal", "Feasible"):
        raise RuntimeError(
            f"No valid schedule found (solver status: {status}). "
            "Try increasing the number of available days."
        )

    assignment = {}
    for i in range(n):
        for d in days:
            if round(pulp.value(x[i, d]) or 0) == 1:
                assignment[counselors[i]] = d
                break

    result = df.copy()
    result["Day Off"] = result["Name"].map(assignment)
    return result.sort_values("Day Off").reset_index(drop=True)


def validate(result: pd.DataFrame):
    issues = []
    for (cabin, day), grp in result[result["Cabin"] != ""].groupby(["Cabin", "Day Off"]):
        if len(grp) > 1:
            issues.append(f"Cabin **{cabin}** — day {day}: {', '.join(grp['Name'])}")
    act_totals = defaultdict(set)
    act_day = defaultdict(lambda: defaultdict(list))
    for _, row in result.iterrows():
        for col in ["Activity1", "Activity2"]:
            if row[col]:
                act_totals[row[col]].add(row["Name"])
                act_day[row[col]][row["Day Off"]].append(row["Name"])
    for act, day_map in act_day.items():
        total = len(act_totals[act])
        if total <= 1:
            continue
        for day, names in day_map.items():
            if len(names) > total / 2:
                issues.append(f"Activity **{act}** — {len(names)}/{total} off on day {day}")
    lead_totals = defaultdict(set)
    lead_day = defaultdict(lambda: defaultdict(list))
    for _, row in result.iterrows():
        if row["Leadership"]:
            lead_totals[row["Leadership"]].add(row["Name"])
            lead_day[row["Leadership"]][row["Day Off"]].append(row["Name"])
    for role, day_map in lead_day.items():
        total = len(lead_totals[role])
        if total <= 1:
            continue
        for day, names in day_map.items():
            if len(names) > total / 2:
                issues.append(f"Leadership **{role}** — {len(names)}/{total} off on day {day}")
    return issues


def colour_row(row, palette):
    day = row["Day Off"]
    if pd.isna(day):
        return [""] * len(row)
    colour = palette.get(int(day), "#FFFFFF")
    r = int(colour[1:3], 16)
    g = int(colour[3:5], 16)
    b = int(colour[5:7], 16)
    brightness = (r * 299 + g * 587 + b * 114) / 1000
    text = "black" if brightness > 128 else "white"
    return [f"background-color:{colour};color:{text}"] * len(row)


def to_excel(df: pd.DataFrame) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day-Off Schedule"

    days = sorted(df["Day Off"].dropna().unique())
    palette = {int(d): DAY_COLOURS[i % len(DAY_COLOURS)] for i, d in enumerate(days)}

    headers = ["Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age", "Day Off"]
    header_fill = PatternFill("solid", fgColor="2D5A27")
    header_font = Font(bold=True, color="FFFFFF", name="Georgia", size=11)
    thin = Side(style="thin", color="C8B89A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, (_, row) in enumerate(df[headers].iterrows(), 2):
        day_val = row["Day Off"]
        hex_colour = palette.get(int(day_val), "FFFFFF").lstrip("#") if not pd.isna(day_val) else "F5F0E8"
        r = int(hex_colour[0:2], 16)
        g = int(hex_colour[2:4], 16)
        b = int(hex_colour[4:6], 16)
        brightness = (r * 299 + g * 587 + b * 114) / 1000
        text_colour = "000000" if brightness > 128 else "FFFFFF"
        fill = PatternFill("solid", fgColor=hex_colour)
        font = Font(name="Calibri", size=10, color=text_colour)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[h])
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    col_widths = [22, 16, 14, 16, 16, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    num_days = st.slider(
        "Days in schedule window",
        min_value=3, max_value=7, value=5,
        help="How many days the schedule spans (e.g. 5 for one week)."
    )
    days_per = st.slider("Days off per counselor", min_value=1, max_value=3, value=1)
    max_per_day = st.slider("Max counselors off per day", min_value=1, max_value=15, value=6)
    st.divider()

    st.markdown("## 📋 Template")
    st.write("Download the template to see the required column format:")
    template_df = pd.DataFrame(
        TEMPLATE_DATA,
        columns=["Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age"]
    )
    template_bytes = io.BytesIO()
    template_df.to_csv(template_bytes, index=False)
    st.download_button(
        "⬇️ Download template CSV",
        template_bytes.getvalue(),
        file_name="carysbrook_counselor_template.csv",
        mime="text/csv",
        use_container_width=True,
    )

    st.divider()

    st.markdown("**Required columns:**")
    st.code("Name, Cabin, Leadership,\nActivity1, Activity2, Age")
    st.caption("Leave cells blank where not applicable.")


# ── Main ──────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    "Upload your counselor file (.csv or .xlsx)",
    type=["csv", "xlsx"],
    help="File must include the columns listed in the sidebar."
)

df_raw = None
if uploaded:
    try:
        if uploaded.name.endswith(".xlsx"):
            df_raw = pd.read_excel(uploaded)
        else:
            df_raw = pd.read_csv(uploaded)
        missing = REQUIRED_COLS - set(df_raw.columns)
        if missing:
            st.error(f"Missing columns: {', '.join(sorted(missing))}. Please check your file matches the template.")
            df_raw = None
        else:
            col1, col2, col3 = st.columns(3)
            col1.metric("Counselors", len(df_raw))
            col2.metric("Cabins", df_raw["Cabin"].replace("", pd.NA).dropna().nunique())
            col3.metric("Activities", pd.concat([df_raw["Activity1"], df_raw["Activity2"]]).replace("", pd.NA).dropna().nunique())
            st.success(f"✅ **{uploaded.name}** loaded successfully.")
    except Exception as e:
        st.error(f"Could not read file: {e}")
else:
    st.info("👆 Upload your counselor spreadsheet above, or download the template from the sidebar to get started.")

if df_raw is not None:
    with st.expander("Preview uploaded data", expanded=False):
        st.dataframe(df_raw, use_container_width=True)

    st.write("")
    if st.button("🗓️ Generate Schedule", type="primary", use_container_width=True):
        df = normalise(df_raw)
        with st.spinner("Finding the best schedule…"):
            try:
                result = schedule(df, num_days=num_days, days_per=days_per, max_per_day=max_per_day)
            except RuntimeError as e:
                st.error(str(e))
                st.stop()

        st.success("Schedule generated!")

        issues = validate(result)
        if issues:
            with st.expander("⚠️ Constraint warnings", expanded=True):
                for issue in issues:
                    st.warning(issue)
        else:
            st.info("✅ All cabin, activity, and leadership constraints satisfied.")

        days_sorted = sorted(result["Day Off"].dropna().unique())
        palette = {int(d): DAY_COLOURS[i % len(DAY_COLOURS)] for i, d in enumerate(days_sorted)}

        st.markdown("### 📊 Counselors off per day")
        counts = result.groupby("Day Off")["Name"].count().reset_index()
        counts.columns = ["Day", "Counselors Off"]
        st.bar_chart(counts.set_index("Day"), color="#2D5A27")

        st.markdown("### 📅 Full Schedule")
        display_cols = ["Name", "Cabin", "Leadership", "Activity1", "Activity2", "Age", "Day Off"]
        styled = (
            result[display_cols]
            .style
            .apply(colour_row, palette=palette, axis=1)
            .set_properties(**{"font-size": "13px"})
        )
        st.dataframe(styled, use_container_width=True, height=600)

        st.markdown("### ⬇️ Download")
        col1, col2 = st.columns(2)
        with col1:
            csv_bytes = result[display_cols].to_csv(index=False).encode()
            st.download_button(
                "Download as CSV",
                csv_bytes,
                file_name="carysbrook_day_off_schedule.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with col2:
            try:
                xl_bytes = to_excel(result)
                st.download_button(
                    "Download as Excel (colour-coded)",
                    xl_bytes,
                    file_name="carysbrook_day_off_schedule.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except ImportError:
                st.warning("Install `openpyxl` for Excel export.")

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="cc-footer">
  Camp Carysbrook &nbsp;·&nbsp; 3500 Camp Carysbrook Road, Riner, VA 24149 &nbsp;·&nbsp; Est. 1923
</div>
""", unsafe_allow_html=True)
